# MultiLingual xml generator from Excel
# pip install openpyxl --->  install this module

import os
import sys
reload(sys)
sys.setdefaultencoding('utf8')

import openpyxl 
wb = openpyxl.load_workbook('Sunbird App Multilingual Strings Sheet.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

blacklist = ['\n','"']

f1=open("Strings.js", 'w+')
print >> f1, "const objectAssign = require('object-assign');"
print >> f1, "const stringsRes = {"
for k in range(2 , 15):
	print >> f1, '\t"'+sheet.cell(row=2, column=k).value+'"'+" : {"
	for l in range(3, 2000):
		all_filled = True
		for x in range(2,14):
			if not sheet.cell(row=l, column=x).value:
				all_filled = False
		if(True):
			key = sheet.cell(row=l, column=1).value
			val = sheet.cell(row=l, column=k).value
			if key is not None  and val is not None :
				for bl in blacklist:
					val=val.replace(bl,"")
				print >> f1, "\t\t"+key+' : "'+val+'",'
			else :
				if key is None and val is not None:
					print "Key not found for value : " + val;
				if(key is not None and val is None):
					print "Value not found for Key : " + key + " for language " + sheet.cell(row=2, column=k).value;
	print >> f1, "\t},"  
print >> f1, "}\n"
print >> f1, "var decideString = function(){"
print >> f1, '\tvar val = window.__CurrentLanguage;'
print >> f1, "\tvar merged = {};"
print >> f1, '\treturn objectAssign({},merged,stringsRes["en_US"],stringsRes[val]);'
print >> f1, "}\n\n"

print >> f1, "var setLanguage = function(lang){"
print >> f1, '\twindow.__CurrentLanguage = lang;'
print >> f1, '\tJBridge.setKey("languagePref",lang);'
print >> f1, "}\n\n"

print >> f1, "exports.setLanguage = setLanguage;"
print >> f1, "exports.strings = decideString;"
print >> f1, "exports.stringsRes = stringsRes;"
